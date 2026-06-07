import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = {
"2025-26": [("Mikal Bridges",82,2692,32.8,0),("Jalen Brunson",74,2590,35.0,0),("Karl-Anthony Towns",75,2322,31.0,0),("OG Anunoby",67,2224,33.2,0),("Josh Hart",66,1994,30.2,0),("Jordan Clarkson",72,1279,17.8,0),("Mitchell Robinson",60,1175,19.6,0),("Landry Shamet",51,1171,23.0,0),("Miles McBride",41,1080,26.3,0),("Tyler Kolek",62,727,11.7,0)],
"2024-25": [("Mikal Bridges",82,3036,37.0,0),("Josh Hart",77,2897,37.6,0),("OG Anunoby",74,2706,36.6,0),("Karl-Anthony Towns",72,2517,35.0,0),("Jalen Brunson",65,2301,35.4,0),("Miles McBride",64,1593,24.9,0),("Precious Achiuwa",57,1170,20.5,0),("Cameron Payne",72,1090,15.1,0),("Landry Shamet",50,762,15.2,0),("Jericho Sims",39,422,10.8,0)],
"2023-24": [("Jalen Brunson",77,2726,35.4,0),("Josh Hart",81,2707,33.4,0),("Donte DiVincenzo",81,2360,29.1,0),("Isaiah Hartenstein",75,1896,25.3,0),("Julius Randle",46,1630,35.4,0),("Miles McBride",68,1328,19.5,0),("Precious Achiuwa",49,1187,24.2,0),("Quentin Grimes",45,910,20.2,0),("OG Anunoby",23,802,34.9,0),("Mitchell Robinson",31,768,24.8,0)],
"2022-23": [("Julius Randle",77,2737,35.5,0),("RJ Barrett",73,2475,33.9,0),("Jalen Brunson",68,2379,35.0,0),("Immanuel Quickley",81,2344,28.9,0),("Quentin Grimes",71,2121,29.9,0),("Isaiah Hartenstein",82,1626,19.8,0),("Mitchell Robinson",59,1591,27.0,0),("Obi Toppin",67,1050,15.7,0),("Jericho Sims",52,812,15.6,0),("Miles McBride",64,760,11.9,0)],
"2021-22": [("Julius Randle",72,2544,35.3,0),("RJ Barrett",70,2417,34.5,0),("Evan Fournier",80,2358,29.5,0),("Alec Burks",81,2318,28.6,0),("Mitchell Robinson",72,1848,25.7,0),("Immanuel Quickley",78,1802,23.1,0),("Obi Toppin",72,1230,17.1,0),("Kemba Walker",37,948,25.6,0),("Taj Gibson",52,946,18.2,0),("Quentin Grimes",46,786,17.1,1)],
"2020-21": [("Julius Randle",71,2667,37.6,0),("RJ Barrett",72,2511,34.9,0),("Reggie Bullock",65,1949,30.0,0),("Nerlens Noel",64,1547,24.2,0),("Elfrid Payton",63,1484,23.6,0),("Alec Burks",49,1255,25.6,0),("Immanuel Quickley",64,1243,19.4,1),("Derrick Rose",35,937,26.8,0),("Taj Gibson",45,936,20.8,0),("Mitchell Robinson",31,853,27.5,0)],
"2019-20": [("Julius Randle",64,2080,32.5,0),("RJ Barrett",56,1704,30.4,1),("Mitchell Robinson",61,1412,23.1,0),("Bobby Portis",66,1393,21.1,0),("Marcus Morris",43,1387,32.3,0),("Elfrid Payton",45,1246,27.7,0),("Frank Ntilikina",57,1187,20.8,0),("Kevin Knox",65,1166,17.9,0),("Taj Gibson",62,1025,16.5,0),("Damyean Dotson",48,836,17.4,0)],
"2018-19": [("Kevin Knox",75,2158,28.8,1),("Damyean Dotson",73,2004,27.5,0),("Noah Vonleh",68,1722,25.3,0),("Emmanuel Mudiay",59,1607,27.2,0),("Tim Hardaway Jr.",46,1499,32.6,0),("Allonzo Trier",64,1459,22.8,1),("Mitchell Robinson",66,1360,20.6,1),("Mario Hezonja",58,1206,20.8,0),("Enes Kanter",44,1128,25.6,0),("Frank Ntilikina",43,904,21.0,0)],
"2017-18": [("Tim Hardaway Jr.",57,1885,33.1,0),("Enes Kanter",71,1830,25.8,0),("Frank Ntilikina",78,1705,21.9,1),("Michael Beasley",74,1653,22.3,0),("Kristaps Porzingis",48,1553,32.4,0),("Jarrett Jack",62,1548,25.0,0),("Kyle O'Quinn",77,1386,18.0,0),("Lance Thomas",73,1353,18.5,0),("Doug McDermott",55,1172,21.3,0),("Trey Burke",36,785,21.8,0)],
"2016-17": [("Carmelo Anthony",74,2538,34.3,0),("Courtney Lee",77,2459,31.9,0),("Kristaps Porzingis",66,2164,32.8,0),("Derrick Rose",64,2082,32.5,0),("Justin Holiday",82,1639,20.0,0),("Brandon Jennings",58,1428,24.6,0),("Willy Hernangomez",72,1324,18.4,1),("Kyle O'Quinn",79,1229,15.6,0),("Mindaugas Kuzminskas",68,1016,14.9,1),("Joakim Noah",46,1015,22.1,0)],
"2015-16": [("Carmelo Anthony",72,2530,35.1,0),("Arron Afflalo",71,2371,33.4,0),("Robin Lopez",82,2219,27.1,0),("Kristaps Porzingis",72,2047,28.4,1),("Langston Galloway",82,2033,24.8,0),("Jose Calderon",72,2024,28.1,0),("Derrick Williams",80,1435,17.9,0),("Lance Thomas",59,1313,22.3,0),("Jerian Grant",76,1265,16.6,1),("Sasha Vujacic",61,908,14.9,0)],
"2014-15": [("Shane Larkin",76,1865,24.5,0),("Jason Smith",82,1785,21.8,0),("Tim Hardaway Jr.",70,1681,24.0,0),("Langston Galloway",45,1457,32.4,1),("Carmelo Anthony",40,1428,35.7,0),("Quincy Acy",68,1287,18.9,0),("Jose Calderon",42,1270,30.2,0),("Lance Thomas",40,1040,26.0,0),("Cole Aldrich",61,976,16.0,0),("Amar'e Stoudemire",36,865,24.0,0)],
"2013-14": [("Carmelo Anthony",77,2982,38.7,0),("JR Smith",74,2421,32.7,0),("Raymond Felton",65,2017,31.0,0),("Iman Shumpert",74,1962,26.5,0),("Tim Hardaway Jr.",81,1875,23.1,1),("Tyson Chandler",55,1662,30.2,0),("Amar'e Stoudemire",65,1466,22.6,0),("Pablo Prigioni",66,1283,19.4,0),("Andrea Bargnani",42,1257,29.9,0),("Kenyon Martin",32,633,19.8,0)],
"2012-13": [("JR Smith",80,2678,33.5,0),("Carmelo Anthony",67,2482,37.0,0),("Raymond Felton",68,2313,34.0,0),("Tyson Chandler",66,2164,32.8,0),("Jason Kidd",76,2043,26.9,0),("Steve Novak",81,1641,20.3,0),("Pablo Prigioni",78,1263,16.2,1),("Iman Shumpert",45,996,22.1,0),("Chris Copeland",56,862,15.4,1),("Ronnie Brewer",46,711,15.5,0)],
"2011-12": [("Tyson Chandler",62,2061,33.2,0),("Landry Fields",66,1894,28.7,0),("Carmelo Anthony",55,1876,34.1,0),("Iman Shumpert",59,1705,28.9,1),("Amar'e Stoudemire",47,1543,32.8,0),("Steve Novak",54,1020,18.9,0),("JR Smith",35,967,27.6,0),("Jeremy Lin",35,940,26.9,0),("Jared Jeffries",39,729,18.7,0),("Toney Douglas",38,656,17.3,0)],
"2010-11": [("Amar'e Stoudemire",78,2870,36.8,0),("Landry Fields",82,2541,31.0,1),("Raymond Felton",54,2074,38.4,0),("Toney Douglas",81,1971,24.3,0),("Wilson Chandler",51,1759,34.5,0),("Danilo Gallinari",48,1671,34.8,0),("Shawne Williams",64,1323,20.7,0),("Ronny Turiaf",64,1141,17.8,0),("Carmelo Anthony",27,977,36.2,0),("Henry Walker",61,784,12.9,0)],
}
coaches = {
"2025-26":"Mike Brown","2024-25":"Tom Thibodeau","2023-24":"Tom Thibodeau","2022-23":"Tom Thibodeau",
"2021-22":"Tom Thibodeau","2020-21":"Tom Thibodeau","2019-20":"David Fizdale / Mike Miller (int.)",
"2018-19":"David Fizdale","2017-18":"Jeff Hornacek","2016-17":"Jeff Hornacek",
"2015-16":"Derek Fisher / Kurt Rambis (int.)","2014-15":"Derek Fisher",
"2013-14":"Mike Woodson","2012-13":"Mike Woodson",
"2011-12":"Mike D'Antoni / Mike Woodson (int.)","2010-11":"Mike D'Antoni",
}
draft = {
("2014-15","Langston Galloway"):"Undrafted (2014)",
("2015-16","Kristaps Porzingis"):"#4 (2015)",
("2015-16","Jerian Grant"):"#19 (2015)",
("2016-17","Willy Hernangomez"):"#35 (2015)",
("2016-17","Mindaugas Kuzminskas"):"Undrafted (2016)",
("2017-18","Frank Ntilikina"):"#8 (2017)",
("2018-19","Kevin Knox"):"#9 (2018)",
("2018-19","Allonzo Trier"):"Undrafted (2018)",
("2018-19","Mitchell Robinson"):"#36 (2018)",
("2019-20","RJ Barrett"):"#3 (2019)",
("2020-21","Immanuel Quickley"):"#25 (2020)",
("2021-22","Quentin Grimes"):"#25 (2021)",
("2013-14","Tim Hardaway Jr."):"#24 (2013)",
("2012-13","Pablo Prigioni"):"Undrafted (2012)",
("2012-13","Chris Copeland"):"Undrafted (2012)",
("2011-12","Iman Shumpert"):"#17 (2011)",
("2010-11","Landry Fields"):"#39 (2010)",
}
wb = Workbook(); FONT="Arial"
hdr_fill=PatternFill("solid",start_color="14213D"); hdr_font=Font(name=FONT,bold=True,color="FFFFFF",size=11)
season_font=Font(name=FONT,bold=True,color="FFFFFF",size=12)
season_fill=PatternFill("solid",start_color="F76900"); rookie_fill=PatternFill("solid",start_color="FFF2CC")
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center"); left=Alignment(horizontal="left",vertical="center")

ws=wb.active; ws.title="By Season"; ws.sheet_view.showGridLines=False
headers=["Rank","Player","Rookie","Draft Pick","GP","Total Min","MPG"]; NC=len(headers); r=1
for season,rows in data.items():
    cell=ws.cell(r,1,f"{season}     Head Coach: {coaches[season]}"); cell.font=season_font
    for c in range(1,NC+1): ws.cell(r,c).fill=season_fill
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NC); cell.alignment=left; ws.row_dimensions[r].height=20; r+=1
    for c,h in enumerate(headers,1):
        cc=ws.cell(r,c,h); cc.font=hdr_font; cc.fill=hdr_fill; cc.alignment=center; cc.border=border
    r+=1
    for i,(player,gp,mins,mpg,rk) in enumerate(rows,1):
        pick=draft.get((season,player),"") if rk else ""
        ws.cell(r,1,i).alignment=center; ws.cell(r,2,player); ws.cell(r,3,"★" if rk else "").alignment=center
        ws.cell(r,4,pick).alignment=center; ws.cell(r,5,gp).alignment=center
        ws.cell(r,6,mins).number_format="#,##0"; ws.cell(r,6).alignment=center
        ws.cell(r,7,mpg).number_format="0.0"; ws.cell(r,7).alignment=center
        for c in range(1,NC+1):
            ws.cell(r,c).border=border
            if not ws.cell(r,c).font.bold: ws.cell(r,c).font=Font(name=FONT,size=11)
            if rk: ws.cell(r,c).fill=rookie_fill
        r+=1
    r+=1
for col,w in zip("ABCDEFG",[7,22,8,13,6,11,7]): ws.column_dimensions[col].width=w
ws.freeze_panes="A2"

ws2=wb.create_sheet("Flat Data"); ws2.sheet_view.showGridLines=False
fh=["Season","Head Coach","Rank","Player","Rookie","Draft Pick","GP","Total Min","MPG"]
for c,h in enumerate(fh,1):
    cc=ws2.cell(1,c,h); cc.font=hdr_font; cc.fill=hdr_fill; cc.alignment=center; cc.border=border
rr=2
for season,rows in data.items():
    for i,(player,gp,mins,mpg,rk) in enumerate(rows,1):
        pick=draft.get((season,player),"") if rk else ""
        ws2.cell(rr,1,season); ws2.cell(rr,2,coaches[season]); ws2.cell(rr,3,i).alignment=center
        ws2.cell(rr,4,player); ws2.cell(rr,5,"Yes" if rk else "").alignment=center; ws2.cell(rr,6,pick).alignment=center
        ws2.cell(rr,7,gp).alignment=center; ws2.cell(rr,8,mins).number_format="#,##0"; ws2.cell(rr,8).alignment=center
        ws2.cell(rr,9,mpg).number_format="0.0"; ws2.cell(rr,9).alignment=center
        for c in range(1,10):
            ws2.cell(rr,c).border=border
            if not ws2.cell(rr,c).font.bold: ws2.cell(rr,c).font=Font(name=FONT,size=11)
            if rk: ws2.cell(rr,c).fill=rookie_fill
        rr+=1
for col,w in zip("ABCDEFGHI",[10,28,7,22,8,13,6,11,7]): ws2.column_dimensions[col].width=w
ws2.auto_filter.ref=f"A1:I{rr-1}"; ws2.freeze_panes="A2"
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "knicks_top10_minutes_2010-2026.xlsx")
wb.save(out_path); print("saved ->", out_path)
